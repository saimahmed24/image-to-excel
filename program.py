import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenPyXLImage
import os
import xlwings as xw

def resize_image_to_cell(image_path, cell_width, cell_height, output_path):
    with Image.open(image_path) as img:
        # Resize image to fit cell height while maintaining aspect ratio
        aspect_ratio = img.width / img.height
        new_height = cell_height
        new_width = int(cell_height * aspect_ratio)

        # If new width exceeds cell width, adjust dimensions
        if new_width > cell_width:
            new_width = cell_width
            new_height = int(cell_width / aspect_ratio)

        img = img.resize((new_width, new_height), Image.LANCZOS)
        img.save(output_path)
        return new_width, new_height

def paste_images_in_excel(folder_path, output_excel, start_cell, placement_type):
    if os.path.exists(output_excel):
        try:
            os.rename(output_excel, output_excel)
        except OSError as e:
            messagebox.showerror("Error", f"{e.filename} - {e.strerror}. Please close the file and try again.")
            return

    if os.path.exists(output_excel):
        workbook = load_workbook(output_excel)
    else:
        workbook = Workbook()

    sheet = workbook.active

    images = sorted([f for f in os.listdir(folder_path) 
                     if f.lower().endswith(('png', 'jpg', 'jpeg', 'gif')) 
                     and not f.startswith('resized_')])

    total_images = len(images)
    print(f"Found {total_images} images")

    if total_images == 0:
        messagebox.showerror("Error", "No images found in the selected folder.")
        return

    app = xw.App(visible=False)
    wb = xw.Book(output_excel)
    sheet_xw = wb.sheets[0]
    app.quit()

    row, col = int(start_cell[1:]), ord(start_cell[0].upper()) - 64

    for i, image_file in enumerate(images):
        image_path = os.path.join(folder_path, image_file)
        resized_image_path = os.path.join(folder_path, f'resized_{image_file}')

        print(f"Processing image: {image_file}")
        print(f"Resizing image: {image_path}")
        print(f"Saving resized image to: {resized_image_path}")

        img_width, img_height = resize_image_to_cell(image_path, 100, 100, resized_image_path)

        current_cell = f'{chr(col + 64)}{row}'
        print(f"Placing image in cell: {current_cell}")

        img = OpenPyXLImage(resized_image_path)
        img.width = img_width
        img.height = img_height

        if placement_type == "Row-wise":
            sheet.column_dimensions[chr(col + 64)].width = img_width / 7  
            sheet.row_dimensions[row].height = img_height * 0.75
            img.anchor = current_cell
            sheet.add_image(img)
            row += 1
        else:
            sheet.column_dimensions[chr(col + 64)].width = img_width / 7  
            sheet.row_dimensions[row].height = img_height * 0.75
            img.anchor = current_cell
            sheet.add_image(img)
            col += 1

    workbook.save(output_excel)
    messagebox.showinfo("Success", f"Images have been pasted in {output_excel}.")

def browse_folder():
    folder_path = filedialog.askdirectory()
    if folder_path:
        folder_entry.config(fg='black')
        folder_entry.delete(0, tk.END)
        folder_entry.insert(0, folder_path)

def browse_file():
    file_path = filedialog.askopenfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        file_entry.config(fg='black')
        file_entry.delete(0, tk.END)
        file_entry.insert(0, file_path)

def start_process():
    folder_path = folder_entry.get()
    output_excel = file_entry.get()
    start_cell = start_cell_entry.get()
    placement_type = placement_type_var.get()
    
    if not folder_path or not output_excel or not start_cell or not placement_type:
        messagebox.showerror("Error", "Please provide all required inputs.")
        return

    paste_images_in_excel(folder_path, output_excel, start_cell, placement_type)

def on_entry_click(event, entry, default_text):
    if entry.get() == default_text:
        entry.delete(0, "end")
        entry.insert(0, '')
        entry.config(fg='black')

def on_focusout(event, entry, default_text):
    if entry.get() == '':
        entry.insert(0, default_text)
        entry.config(fg='grey')

# GUI Setup
root = tk.Tk()
root.title("Image to Excel")

# Set the window size
window_width = 540
window_height = 250

# Get the screen width and height
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# Calculate position x, y
x = (screen_width // 2) - (window_width // 2)
y = (screen_height // 2) - (window_height // 2)

# Set the window position and size
root.geometry(f"{window_width}x{window_height}+{x}+{y}")

# Configure the background color and layout
root.configure(bg="#1f1f1f")

# Folder Path
default_folder_text = "Select folder..."
folder_entry = tk.Entry(root, width=50, fg='grey')
folder_entry.insert(0, default_folder_text)
folder_entry.bind('<FocusIn>', lambda event: on_entry_click(event, folder_entry, default_folder_text))
folder_entry.bind('<FocusOut>', lambda event: on_focusout(event, folder_entry, default_folder_text))
folder_entry.grid(row=0, column=1, padx=10, pady=10)
tk.Button(root, text="Browse", command=browse_folder, bg="#1f1f1f", fg="#ffffff").grid(row=0, column=2, padx=10, pady=10)
tk.Label(root, text="Folder Path:", bg="#1f1f1f", fg="#ffffff").grid(row=0, column=0, padx=10, pady=10, sticky=tk.W)

# Excel File
default_file_text = "Select Excel file..."
file_entry = tk.Entry(root, width=50, fg='grey')
file_entry.insert(0, default_file_text)
file_entry.bind('<FocusIn>', lambda event: on_entry_click(event, file_entry, default_file_text))
file_entry.bind('<FocusOut>', lambda event: on_focusout(event, file_entry, default_file_text))
file_entry.grid(row=1, column=1, padx=10, pady=10)
tk.Button(root, text="Browse", command=browse_file, bg="#1f1f1f", fg="#ffffff").grid(row=1, column=2, padx=10, pady=10)
tk.Label(root, text="Excel File:", bg="#1f1f1f", fg="#ffffff").grid(row=1, column=0, padx=10, pady=10, sticky=tk.W)

# Start Cell
tk.Label(root, text="Start Cell (e.g., A1):", bg="#1f1f1f", fg="#ffffff").grid(row=2, column=0, padx=10, pady=10, sticky=tk.W)
start_cell_entry = tk.Entry(root, width=10)
start_cell_entry.grid(row=2, column=1, padx=10, pady=10, sticky=tk.W)

# Placement Type
tk.Label(root, text="Placement Type:", bg="#1f1f1f", fg="#ffffff").grid(row=3, column=0, padx=10, pady=10, sticky=tk.W)
placement_type_var = tk.StringVar()
placement_type_dropdown = ttk.Combobox(root, textvariable=placement_type_var, values=["Row-wise", "Column-wise"], state="readonly", width=10)
placement_type_dropdown.grid(row=3, column=1, padx=10, pady=10, sticky=tk.W)
placement_type_dropdown.current(0)

# Start Button
tk.Button(root, text="Start", command=start_process, bg="#1f1f1f", fg="#ffffff").grid(row=4, column=1, padx=10, pady=20)

root.mainloop()
